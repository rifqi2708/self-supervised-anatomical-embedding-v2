"""Configuration, cohort discovery, and manifest helpers."""

from __future__ import annotations

import hashlib
import json
import os
import re
import subprocess
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

import yaml

MANIFEST_SCHEMA_VERSION = 1
DEFAULT_SUBJECT_START = 21
DEFAULT_SUBJECT_END = 48
DEFAULT_RUN_ID = "totalseg-2.16.0-organs-v1"
DEFAULT_REGISTRY = Path(__file__).with_name("organs.yaml")
VALID_SEXES = {"M", "F"}


class WorkflowError(RuntimeError):
    """Raised when workflow configuration or source data is invalid."""


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def sha256_file(path: Path, chunk_size: int = 8 * 1024 * 1024) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while True:
            chunk = handle.read(chunk_size)
            if not chunk:
                break
            digest.update(chunk)
    return digest.hexdigest()


def atomic_write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{path.name}.", suffix=".tmp", dir=str(path.parent)
    )
    try:
        with os.fdopen(descriptor, "w", encoding="utf-8") as handle:
            json.dump(value, handle, indent=2, sort_keys=True)
            handle.write("\n")
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary_name, path)
    except BaseException:
        try:
            os.unlink(temporary_name)
        except FileNotFoundError:
            pass
        raise


def canonical_subject_id(value: object) -> str:
    if isinstance(value, bool):
        raise ValueError(f"Invalid subject identifier: {value!r}")
    if isinstance(value, (int, float)) and float(value).is_integer():
        number = int(value)
    else:
        match = re.search(r"(\d{1,3})\s*$", str(value).strip())
        if not match:
            raise ValueError(f"Invalid subject identifier: {value!r}")
        number = int(match.group(1))
    return f"quadra_hc_{number:03d}"


def subject_number(subject_id: str) -> int:
    return int(canonical_subject_id(subject_id).rsplit("_", 1)[1])


def load_registry(path: Path | str = DEFAULT_REGISTRY) -> dict[str, Any]:
    registry_path = Path(path).expanduser().resolve()
    with registry_path.open("r", encoding="utf-8") as handle:
        registry = yaml.safe_load(handle)
    if not isinstance(registry, dict):
        raise WorkflowError("Organ registry must contain a YAML mapping")
    if int(registry.get("schema_version", -1)) != 1:
        raise WorkflowError("Unsupported organ registry schema")
    tasks = registry.get("tasks")
    organs = registry.get("organs")
    derived = registry.get("derived_organs")
    intermediates = registry.get("intermediates")
    if not isinstance(tasks, dict) or not tasks:
        raise WorkflowError("Organ registry must define tasks")
    if not isinstance(organs, list) or not organs:
        raise WorkflowError("Organ registry must define organs")
    if not isinstance(derived, list) or not isinstance(intermediates, list):
        raise WorkflowError("Organ registry must define derived_organs and intermediates")

    filenames: list[str] = []
    for entry in organs:
        required = {"display_name", "filename", "task", "source_class"}
        if not isinstance(entry, dict) or not required.issubset(entry):
            raise WorkflowError(f"Invalid organ entry: {entry!r}")
        if entry["task"] not in tasks:
            raise WorkflowError(f"Unknown task for {entry['filename']}: {entry['task']}")
        sexes = set(entry.get("sexes", VALID_SEXES))
        if not sexes or not sexes.issubset(VALID_SEXES):
            raise WorkflowError(f"Invalid sex rule for {entry['filename']}")
        filenames.append(str(entry["filename"]))
    for entry in derived:
        if not isinstance(entry, dict) or not {
            "display_name",
            "filename",
            "derivation",
            "region",
        }.issubset(entry):
            raise WorkflowError(f"Invalid derived organ entry: {entry!r}")
        filenames.append(str(entry["filename"]))
    if len(filenames) != len(set(filenames)):
        raise WorkflowError("Final organ filenames must be unique")
    for entry in intermediates:
        if not isinstance(entry, dict) or not {"filename", "task", "source_class"}.issubset(entry):
            raise WorkflowError(f"Invalid intermediate entry: {entry!r}")
        if entry["task"] not in tasks:
            raise WorkflowError(f"Unknown intermediate task: {entry['task']}")
    return registry


def registry_identity(path: Path | str) -> dict[str, str]:
    resolved = Path(path).expanduser().resolve()
    return {"path": str(resolved), "sha256": sha256_file(resolved)}


def expected_mask_names(registry: dict[str, Any], sex: str) -> list[str]:
    normalized = str(sex).strip().upper()
    if normalized not in VALID_SEXES:
        raise WorkflowError(f"Unknown sex: {sex!r}")
    direct = [
        str(entry["filename"])
        for entry in registry["organs"]
        if normalized in set(entry.get("sexes", VALID_SEXES))
    ]
    derived = [str(entry["filename"]) for entry in registry["derived_organs"]]
    return direct + derived


def task_classes(registry: dict[str, Any], sex: str) -> dict[str, list[str]]:
    selected: dict[str, set[str]] = {task: set() for task in registry["tasks"]}
    for entry in registry["organs"]:
        if sex in set(entry.get("sexes", VALID_SEXES)):
            selected[str(entry["task"])].add(str(entry["source_class"]))
    for entry in registry["intermediates"]:
        selected[str(entry["task"])].add(str(entry["source_class"]))
    return {task: sorted(classes) for task, classes in selected.items() if classes}


def _load_demographics_xlsx(path: Path, sheet_name: str) -> list[tuple[object, object]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise WorkflowError(
            "Reading .xlsx demographics requires openpyxl; install the dedicated requirements"
        ) from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise WorkflowError(
                f"Demographics sheet {sheet_name!r} not found; available: {workbook.sheetnames}"
            )
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    header_index = None
    subject_column = sex_column = None
    for index, row in enumerate(rows[:10]):
        labels = [str(value).strip().lower() if value is not None else "" for value in row]
        if "subject" in labels and "sex" in labels:
            header_index = index
            subject_column = labels.index("subject")
            sex_column = labels.index("sex")
            break
    if header_index is None or subject_column is None or sex_column is None:
        raise WorkflowError("Could not locate Subject and Sex columns in demographics workbook")
    return [
        (
            row[subject_column] if subject_column < len(row) else None,
            row[sex_column] if sex_column < len(row) else None,
        )
        for row in rows[header_index + 1 :]
    ]


def _load_demographics_csv(path: Path) -> list[tuple[object, object]]:
    import csv

    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if not reader.fieldnames:
            raise WorkflowError("Demographics CSV is empty")
        normalized = {name.strip().lower(): name for name in reader.fieldnames}
        if "subject" not in normalized or "sex" not in normalized:
            raise WorkflowError("Demographics CSV must contain Subject and Sex columns")
        return [
            (row[normalized["subject"]], row[normalized["sex"]])
            for row in reader
        ]


def read_demographics(
    path: Path | str,
    subject_ids: Iterable[str],
    sheet_name: str = "Demographics (All)",
) -> dict[str, str]:
    demographics_path = Path(path).expanduser().resolve()
    wanted = {canonical_subject_id(subject) for subject in subject_ids}
    if demographics_path.suffix.lower() == ".xlsx":
        raw_rows = _load_demographics_xlsx(demographics_path, sheet_name)
    elif demographics_path.suffix.lower() == ".csv":
        raw_rows = _load_demographics_csv(demographics_path)
    else:
        raise WorkflowError("Demographics must be an .xlsx or .csv file")

    result: dict[str, str] = {}
    duplicates: set[str] = set()
    for raw_subject, raw_sex in raw_rows:
        if raw_subject is None or str(raw_subject).strip() == "":
            continue
        try:
            subject = canonical_subject_id(raw_subject)
        except ValueError:
            continue
        if subject not in wanted:
            continue
        sex = str(raw_sex).strip().upper()
        if sex not in VALID_SEXES:
            raise WorkflowError(f"Invalid sex {raw_sex!r} for {subject}")
        if subject in result:
            duplicates.add(subject)
        result[subject] = sex
    if duplicates:
        raise WorkflowError(f"Duplicate demographic rows: {', '.join(sorted(duplicates))}")
    missing = sorted(wanted - set(result))
    if missing:
        raise WorkflowError(f"Missing demographics: {', '.join(missing)}")
    return result


def discover_scans(
    dataset_root: Path | str,
    subject_start: int = DEFAULT_SUBJECT_START,
    subject_end: int = DEFAULT_SUBJECT_END,
) -> list[dict[str, str]]:
    if subject_start < DEFAULT_SUBJECT_START or subject_end > DEFAULT_SUBJECT_END:
        raise WorkflowError(
            "Evaluation cohort is restricted to subjects 021-048; subjects 001-020 are reserved"
        )
    if subject_start > subject_end:
        raise WorkflowError("subject_start must not exceed subject_end")
    root = Path(dataset_root).expanduser().resolve()
    scans: list[dict[str, str]] = []
    missing: list[str] = []
    for number in range(subject_start, subject_end + 1):
        subject = f"quadra_hc_{number:03d}"
        subject_directory = root / subject.upper()
        for session in ("test", "retest"):
            path = subject_directory / f"{session}_CT-AC.nii.gz"
            if not path.is_file():
                missing.append(str(path))
            else:
                scans.append(
                    {
                        "subject_id": subject,
                        "session": session,
                        "input_path": str(path),
                    }
                )
    if missing:
        preview = "\n".join(missing[:10])
        raise WorkflowError(f"Missing {len(missing)} expected CT files:\n{preview}")
    return scans


def git_commit(project_root: Path) -> str | None:
    try:
        completed = subprocess.run(
            ["git", "rev-parse", "HEAD"],
            cwd=project_root,
            check=True,
            capture_output=True,
            text=True,
        )
    except (OSError, subprocess.CalledProcessError):
        return None
    return completed.stdout.strip() or None


def prepare_manifest(
    dataset_root: Path | str,
    demographics_path: Path | str,
    registry_path: Path | str = DEFAULT_REGISTRY,
    subject_start: int = DEFAULT_SUBJECT_START,
    subject_end: int = DEFAULT_SUBJECT_END,
    run_id: str = DEFAULT_RUN_ID,
    project_root: Path | None = None,
) -> dict[str, Any]:
    registry = load_registry(registry_path)
    scans = discover_scans(dataset_root, subject_start, subject_end)
    subjects = sorted({scan["subject_id"] for scan in scans})
    demographics = read_demographics(demographics_path, subjects)
    registry_info = registry_identity(registry_path)
    demographics_resolved = Path(demographics_path).expanduser().resolve()

    manifest_scans: list[dict[str, Any]] = []
    for scan in scans:
        input_path = Path(scan["input_path"])
        sex = demographics[scan["subject_id"]]
        manifest_scans.append(
            {
                **scan,
                "sex": sex,
                "input_sha256": sha256_file(input_path),
                "input_size_bytes": input_path.stat().st_size,
                "expected_masks": expected_mask_names(registry, sex),
                "task_classes": task_classes(registry, sex),
            }
        )
    male_scans = sum(scan["sex"] == "M" for scan in manifest_scans)
    female_scans = sum(scan["sex"] == "F" for scan in manifest_scans)
    return {
        "schema_version": MANIFEST_SCHEMA_VERSION,
        "created_at": utc_now(),
        "run_id": run_id,
        "git_commit": git_commit(project_root) if project_root else None,
        "totalsegmentator_version": registry["totalsegmentator_version"],
        "registry": registry_info,
        "demographics": {
            "path": str(demographics_resolved),
            "sha256": sha256_file(demographics_resolved),
            "sheet": "Demographics (All)",
        },
        "dataset_root": str(Path(dataset_root).expanduser().resolve()),
        "subject_range": [subject_start, subject_end],
        "summary": {
            "subjects": len(subjects),
            "scans": len(manifest_scans),
            "male_subjects": sum(demographics[subject] == "M" for subject in subjects),
            "female_subjects": sum(demographics[subject] == "F" for subject in subjects),
            "male_scans": male_scans,
            "female_scans": female_scans,
            "expected_masks": sum(len(scan["expected_masks"]) for scan in manifest_scans),
        },
        "scans": manifest_scans,
    }


def load_manifest(path: Path | str) -> dict[str, Any]:
    manifest_path = Path(path).expanduser().resolve()
    with manifest_path.open("r", encoding="utf-8") as handle:
        manifest = json.load(handle)
    if int(manifest.get("schema_version", -1)) != MANIFEST_SCHEMA_VERSION:
        raise WorkflowError(f"Unsupported manifest schema in {manifest_path}")
    if not isinstance(manifest.get("scans"), list):
        raise WorkflowError(f"Manifest has no scan list: {manifest_path}")
    return manifest


def find_scan(manifest: dict[str, Any], subject: str, session: str) -> dict[str, Any]:
    canonical = canonical_subject_id(subject)
    normalized_session = session.strip().lower()
    if normalized_session not in {"test", "retest"}:
        raise WorkflowError(f"Unknown session: {session!r}")
    matches = [
        scan
        for scan in manifest["scans"]
        if scan["subject_id"] == canonical and scan["session"] == normalized_session
    ]
    if len(matches) != 1:
        raise WorkflowError(f"Expected one manifest scan for {canonical} {normalized_session}")
    return matches[0]
